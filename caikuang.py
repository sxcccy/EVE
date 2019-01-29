# -*- coding: utf8 -*-
import tkinter as tk
import pyautogui as pag
import time
import threading
import openpyxl
from openpyxl.styles import PatternFill, Fill
from PIL import Image

#---------创建界面---------
window=tk.Tk()
window.title('caikuang')
window.geometry('200x300')
window.wm_attributes('-topmost',1)# 顶置窗口


var1 = tk.StringVar()
label1 = tk.Label(window,textvariable=var1,bg='green', font=('Arial', 12), width=15, height=2)
label1.pack()
var2 = tk.StringVar()
label2 = tk.Label(window,textvariable=var2,bg='blue', font=('Arial', 12), width=15, height=2)
label2.pack()
var3 = tk.StringVar()
label3 = tk.Label(window,textvariable=var3,bg='white', font=('Arial', 12), width=15, height=2)
label3.pack()

def fun_timer():
    try:
        x1, x2, y1, y2 = 1484, 1644, 787, 813 #找色区域
        pag.moveTo(1689, 796, 0.5)
        pag.click()
        time.sleep(1)
        img = pag.screenshot()
        break_flag = False
        for ix in range(x1, x2):
            for iy in range(y1, y2):
                (r, g, b) = img.getpixel((ix, iy))
                if g>200:
                    var1.set((r, g, b))
                    pag.press('f',1,1)
                    pag.moveTo(1723, 271, 0.5)
                    pag.click()
                    pag.press('q', 1, 1)
                    time.sleep(0.5)
                    pag.press('ctrlleft', 1, 1)
                    pag.moveTo(1723, 271+19, 0.5)
                    pag.click()
                    pag.press('ctrlleft', 1, 1)
                    pag.moveTo(1723, 271 + 19*2, 0.5)
                    pag.click()
                    pag.press('ctrlleft', 1, 1)
                    break_flag = True
                    break
                elif r>200:
                    var2.set((r, g, b))
                    break_flag = True
                    break
                var3.set('error')
            if break_flag:
                break

        time.sleep(5)
        global timer
        timer = threading.Timer(5, fun_timer)
        timer.start()
    except Exception as e:
        pag.alert(e)


timer = threading.Timer(1, fun_timer)
def srpstart():
    global timer
    timer.start()

def srpend():
    global timer
    timer.cancel()

def test():
    x1,x2,y1,y2=1484,1644,787,813
    pag.moveTo(1689, 796, 0.5)
    pag.click()
    time.sleep(1)
    img=pag.screenshot()
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('test')
    im = Image.new("RGB", (x2-x1, y2-y1))
    for ix in range(x1,x2):
        for iy in range(y1,y2):
            (r, g, b) = img.getpixel((ix, iy))
            im.putpixel((ix-x1, iy-y1), (r, g, b))
            rgba = im.getpixel((ix-x1, iy-y1))
            colorHex = hex(rgba[0])[2:].zfill(2) + hex(rgba[1])[2:].zfill(2) + hex(rgba[2])[2:].zfill(2)
            fill = PatternFill(fill_type='solid', start_color=colorHex, end_color=colorHex)
            ws.cell(row=iy-y1 + 1,column=ix-x1 + 1 ).fill = fill
            ws.cell(row=iy - y1 + 1, column=ix - x1 + 1).value = str(ix)+','+str(iy)+'='+str(r)+','+str(g)+','+str(b)
    wb.save(r'd:\Desktop\test12.xlsx')
    pag.alert('test over!')


btn_start = tk.Button(window,text='Start' , width=15, height=2,command=srpstart)
btn_start.pack()

btn_stop = tk.Button(window,text='Stop', width=15, height=2,command=srpend)
btn_stop.pack()

btn_test = tk.Button(window,text='Test', width=15, height=2,command=test)
btn_test.pack()



window.mainloop()

